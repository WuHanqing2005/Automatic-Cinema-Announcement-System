'''
软件名称：电影院自动广播测试系统（通过爬虫）
版本号：2025.03.06 08:20(正式版)
软件版权归属：吴瀚庆
未经允许，禁止盗用，侵权必究

有意请联系软件作者 吴瀚庆
微信：whq20050121
手机：19528873640
邮箱：m19528873640@outlook.com
欢迎提出宝贵意见，感谢支持！

打包指令：pyinstaller main.spec
'''

import os
import time
import datetime
from datetime import datetime
import pandas as pd
from asyncio import exceptions
import requests
from bs4 import BeautifulSoup
from fake_useragent import UserAgent
import random
import openpyxl  # 导入 openpyxl 库以处理 Excel 文件
import tkinter as tk
from tkinter import ttk, messagebox
import threading
import pygame
from pydub import AudioSegment


film_played = []    # 初始化已播放电影列表

version_code = '2025.03.04 19:30(正式版)'     # 版本号

# 全局变量
cycle_time = 2  # 默认重复播放次数
pre_minute = 5  # 默认提前检票分钟数
table_lock = threading.Lock()   # 线程锁，防止访问表格时出现冲突

# 读取info.txt文件中的配置信息
config = {}
with open('info.txt', 'r', encoding='utf-8') as f:
    for line in f:
        key, value = line.strip().split(' = ', 1)
        config[key] = eval(value)

global uid_list_admin, machine_info, api_url, appToken, appKey, url
uid_list_admin = config['uid_list_admin']
machine_info = config['machine_info']
api_url = config['api_url']     # WXPusher的api_url
appToken = config['appToken']
appKey = config['appKey']
url = config['url']             # 爬取的猫眼url

import fake_useragent

# 设置 fake_useragent 的数据文件路径
current_dir = os.path.dirname(os.path.abspath(__file__))
fake_useragent_path = os.path.join(current_dir, 'packages', 'fake_useragent', 'data')
os.environ['FAKE_USERAGENT_PATH'] = fake_useragent_path

# 初始化 UserAgent
ua = fake_useragent.UserAgent()

# 生成100个随机请求头
ua = UserAgent()
user_agents = [ua.random for _ in range(100)]

# 返回随机请求头的函数
def get_random_user_agent():
    # ua = UserAgent()
    # user_agents = [ua.random for _ in range(100)]
    return random.choice(user_agents)


# 写错误日志的函数
def write_error_log(error_message):
    import datetime
    # 获取当前时间
    now = datetime.datetime.now()
    # 格式化时间为字符串
    timestamp = now.strftime("%Y-%m-%d %H:%M:%S")
    
    # 构造日志信息
    log_message = f"[{timestamp}]\nMachine Info:{machine_info}\nVersion:{version_code}\nERROR: {error_message}\n\n"
    
    # 定义日志文件名
    log_filename = f"error_log.txt"
    
    # 打开文件，写入日志信息
    with open(log_filename, "a") as file:
        file.write(log_message)
    
    # 定义一个函数用于发送错误日志
    def send_error_log():
        try:
            for uid in uid_list_admin:
                # 构建发送消息的请求参数
                data = {
                    "appToken": appToken,
                    "content": str(error_message),  # 使用转换后的字符串
                    "summary": "电影院自动广播测试系统 错误信息",
                    "contentType": 1,
                    "topicIds": [123],
                    "uids": [uid]  # 替换成要发送消息的微信用户的userId
                }

                # 发送POST请求
                response = requests.post(api_url, json=data)

                # 打印返回的结果
                # print(response.json())
        except Exception as e:
            print(f"发送错误日志失败: {e}")

    # 创建一个线程来执行发送错误日志的操作
    send_thread = threading.Thread(target=send_error_log)
    send_thread.start()

    # 设置超时时间为1秒
    send_thread.join(timeout=1)

    # 如果线程仍然在运行（即超时），则弹出错误窗口
    if send_thread.is_alive():
        # 超时处理，显示错误信息
        messagebox.showwarning("警告", "由于网络原因，错误日志未能成功发送。")
        
        # 继续执行主程序
        pass               

# 检查电影名文件缺失的函数
def check_movie_name():
    missing_files = set()  # 使用集合存储缺失的电影名称.wav文件，自动去重
    # 遍历data数组中的每一行数据
    try:
        for row in data:
            film_name = row[0]  # 获取电影名称
            # 构造电影名称.wav文件的路径
            file_path = os.path.join('material', 'filmname_cn', f'{film_name}.wav')
            # 检查文件是否存在
            if not os.path.exists(file_path):
                missing_files.add(file_path)  # 如果文件不存在，添加到集合中

        # 如果有缺失的文件，弹出报错窗口
        if missing_files:
            warning_message = "以下电影名称.wav文件缺失：\n" + "\n".join(missing_files)
            messagebox.showwarning("警告", warning_message)
            write_error_log(warning_message)
        else:
            print("电影名文件缺失检查完毕，所有电影名称wav文件都存在。")
    except:
        warning_message = "由于未知的原因，电影名文件缺失检查未能进行，但是并不影响主程序运行。"
        messagebox.showwarning("警告", warning_message)
        print(warning_message)
        write_error_log(warning_message)



def delete_all_files_in_directory(directory_path):
    # 检查目录是否存在
    if not os.path.exists(directory_path):
        print(f"目录'{directory_path}'不存在。")
        return False

    # 检查路径是否确实是一个目录
    if not os.path.isdir(directory_path):
        print(f"'{directory_path}'不是一个目录。")
        return False

    try:
        # 遍历目录中的所有文件和文件夹
        for filename in os.listdir(directory_path):
            file_path = os.path.join(directory_path, filename)
            
            # 如果是文件，则删除
            if os.path.isfile(file_path) or os.path.islink(file_path):
                os.unlink(file_path)
                print(f"文件'{file_path}'已被删除。")
            # 如果是目录，则递归删除
            elif os.path.isdir(file_path):
                delete_all_files_in_directory(file_path)
                os.rmdir(file_path)
                print(f"文件夹'{file_path}'已被删除。")
    except Exception as e:
        print(f'无法删除{directory_path}。原因: {e}')
        write_error_log(e)
        return False
    return True


def update_progress(step, total, progress_bar, label, text):
    # 更新进度条的值
    progress_bar.configure(value=(step / total) * 100)
    # 更新标签的文本
    label.config(text=f"Importing voice packs:  ({step} / {total})\n{text}")

# 将文件夹中所有子文件夹中的wav文件全部转换为双声道
def convert_to_stereo(input_folder):
    # 初始化Tkinter窗口
    root = tk.Tk()
    root.title("Converting audio files")
    root.geometry("500x120")

    # 创建进度条
    progress_bar = ttk.Progressbar(root, orient="horizontal", length=300, mode="determinate")
    progress_bar.pack(pady=20, padx=20, fill="both", expand=True)

    # 创建一个标签控件
    label = tk.Label(root, text="Converting audio files", font=("Times New Roman", 15))
    label.pack()

    # 获取 material 文件夹下的文件夹个数
    entries = os.listdir(input_folder)
    subfolders = [entry for entry in entries if os.path.isdir(os.path.join(input_folder, entry))]
    total_subfolders = len(subfolders)

    # 循环执行 int(len(subfolders)) 次，也就是要遍历 int(len(subfolders)) 个子文件夹
    i = 0

    # 遍历input_folder中的所有文件和子文件夹
    for current_root, dirs, files in os.walk(input_folder):  # 第一层循环，遍历子文件夹
        for filename in files:  # 第二层循环，遍历子文件夹中的语音包
            if filename.endswith(".wav"):
                try:
                    input_path = os.path.join(current_root, filename)
                    audio = AudioSegment.from_wav(input_path)
                    stereo_audio = audio.set_channels(2)
                    stereo_audio.export(input_path, format="wav")
                except Exception as error_message:
                    messagebox.showerror("Error", f"Failed to convert audio file: {filename}\nError message: {error_message}")
                    write_error_log(f"Failed to convert audio file: {filename}\nError message: {error_message}")
                    continue

        # 更新进度条和标签
        step = i
        total = total_subfolders
        update_progress(step, total, progress_bar, label, current_root)  # 使用current_root而不是root
        root.update()  # 刷新主窗口，以便更新进度条和标签
        i += 1  # 增加当前的步数

    # 在这里不需要调用root.mainloop()，因为程序会在所有文件处理完毕后退出
    # 如果需要保持窗口打开，确保在最后调用root.mainloop()，并且在此之前不要退出程序

    # 运行完毕，弹出提示窗口
    # messagebox.showinfo("Information", "Converting audio files to dual-channel successfully!")
    root.destroy()

# 爬取猫眼网获取电影排期的函数
def fetch_movie_schedules(url, progress_window):
    global successful, cinema_name, cinema_address  # 添加全局变量

    try:
        max_attempts = 50
        attempts = 0
        successful = False
        data = []

        while attempts < max_attempts and not successful:
            headers = {
                'User-Agent': get_random_user_agent()
            }

            try:
                response = requests.get(url, headers=headers)
                response.raise_for_status()

                # 打印状态码和响应内容的前100个字符用于调试
                print(f"Status Code: {response.status_code}")
                print(f"Response Content (First 100 chars): {response.text[:100]}")

                soup = BeautifulSoup(response.text, 'html.parser')

                # 新增：爬取电影院名称和地址
                cinema_info = soup.find('div', class_='cinema-main clearfix')
                if cinema_info:
                    global cinema_name, cinema_address
                    cinema_name = cinema_info.find('h1', class_='name text-ellipsis').text.strip()
                    cinema_address = cinema_info.find('div', class_='address text-ellipsis').text.strip()

                    # 调试代码
                    print(f"Cinema Name: {cinema_name}")
                    print(f"Cinema Address: {cinema_address}")

                movie_sections = soup.find_all('div', class_='show-list')

                if not movie_sections:
                    print("没有找到电影信息，等待1秒后重试...")
                    attempts += 1
                    continue

                for movie_section in movie_sections:
                    movie_name = movie_section.find('h2', class_='movie-name').text.strip()
                    date_sections = movie_section.find('div', class_='show-date')
                
                    if not date_sections:
                        print(f"没有找到日期信息 for {movie_name}")
                        continue

                    date_items = date_sections.find_all('span', class_='date-item')
                    for date_item in date_items:
                        date_text = date_item.text.strip()

                        if " " in date_text:
                            date_split = date_text.split(" ")
                            date_name = date_split[0]
                            date_day = " ".join(date_split[1:])
                        else:
                            date_name = date_text
                            date_day = ""

                        date_index = date_item['data-index']
                        show_list = movie_section.find_all('div', class_='plist-container')
                        active_show_list = show_list[int(date_index)]

                        time_slots = active_show_list.find_all('tr')
                        for time_slot in time_slots:
                            begin_time = time_slot.find('span', class_='begin-time')
                            end_time = time_slot.find('span', class_='end-time')
                            hall = time_slot.find('span', class_='hall')

                            if begin_time and end_time and hall:
                                end_time_text = end_time.text.strip().split('散场')[0].strip()
                                hall_text = hall.text.strip().split('厅')[0] + '厅'  # 提取“厅”之前的内容

                                data.append([
                                    movie_name, 
                                    date_name, 
                                    date_day, 
                                    begin_time.text.strip(), 
                                    end_time_text, 
                                    hall_text
                                ])
                    
                successful = True
                break

            except requests.RequestException as e:
                print(f"请求错误: {e}")
                write_error_log(e)
                attempts += 1
            except Exception as e:
                print(f"发生错误: {e}")
                attempts += 1

        if not successful:
            print("已达到最大尝试次数，未能获取电影信息。")
            return []  # 返回一个空列表

        # 排序数据
        # 自定义排序规则
        def sort_key(row):
            date_day = row[2]  # 取出日期详情
            start_time = row[3]  # 取出开始时间
            # 将日期详情和开始时间组合成一个元组进行排序
            return (date_day, start_time)

        sorted_data = sorted(data, key=sort_key)

        # 调试信息，输出整理后的二维数组前5个元素
        print('The first 5 rows of sorted data:')
        for i in range(5):
            print(sorted_data[i])

        # 将二维数组写入 Excel 文件
        write_to_excel(sorted_data)
    
        # 关闭进度窗口
        progress_window.destroy()
    
        return sorted_data  # 返回排序后的数据
    except Exception as e:
        write_error_log(e)

# 将电影信息data写入Excel文件的函数
def write_to_excel(data):
    try:
        # 尝试获取锁，设置超时时间为1秒
        if table_lock.acquire(timeout=1):
            try:
                # 创建一个新的工作簿和工作表
                workbook = openpyxl.Workbook()
                sheet = workbook.active
                sheet.title = 'Movie Schedules'

                # 写入标题行
                sheet.append(['Filmname', 'Date name', 'Date day', 'Start time', 'End time', 'Hall No.'])

                # 写入每一行数据
                for row in data:
                    sheet.append(row)

                # 保存 Excel 文件
                workbook.save('data.xlsx')
                print("数据成功写入 data.xlsx 文件。")
            finally:
                # 释放锁
                table_lock.release()
    except Exception as e:
        print(f"写入Excel文件错误: {e}")
        write_error_log(e)

# 从xlsx表格读取电影信息并显示的函数
def read_from_excel():
    try:
        # 尝试获取锁，设置超时时间为1秒
        if table_lock.acquire(timeout=1):
            try:
                workbook = openpyxl.load_workbook('data.xlsx')
                sheet = workbook.active
                data = []
                for row in sheet.iter_rows(min_row=2, values_only=True):
                    data.append(row)
                return data
            finally:
                # 释放锁
                table_lock.release()
    except Exception as e:
        print(f"读取Excel文件错误: {e}")
        write_error_log(e)
        return []

# 更新表格显示的函数
def update_table(data):
    try:
        # 尝试获取锁，设置超时时间为1秒
        if table_lock.acquire(timeout=1):
            try:
                for i in reversed(table.get_children()):
                    table.delete(i)
                for row in data:
                    table.insert('', 'end', values=row)
            finally:
                # 释放锁
                table_lock.release()
    except Exception as e:
        print(f"更新表格时发生错误: {e}")
        write_error_log(e)

# 修改电影信息的函数
def modify_movie_info():
    global data
    selection = movie_drop_down.get()
    if not selection:
        messagebox.showwarning("警告", "请选择一部电影")
        return

    # 获取选中的电影信息
    selected_data = None
    for row in data:
        if f"{row[0]}-{row[1]}-{row[3]}" == selection:
            selected_data = row
            break

    if not selected_data:
        messagebox.showerror("错误", "未找到选中的电影信息")
        return

    # 弹出修改窗口
    modify_window = tk.Toplevel(root)
    modify_window.title("修改电影信息")

    # 获取屏幕宽高
    screen_width = modify_window.winfo_screenwidth()
    screen_height = modify_window.winfo_screenheight()

    # 设置窗口大小
    window_width = 350
    window_height = 300

    # 计算居中位置
    x = (screen_width // 2) - (window_width // 2)
    y = (screen_height // 2) - (window_height // 2)

    # 设置窗口大小及位置
    modify_window.geometry(f"{window_width}x{window_height}+{x}+{y}")

    # 创建输入框和标签
    labels = ["电影名称", "日期", "日期详情", "开始时间", "结束时间", "放映厅"]
    entries = []
    for i, label_text in enumerate(labels):
        tk.Label(modify_window, text=label_text).grid(row=i, column=0, padx=10, pady=5)
        entry = tk.Entry(modify_window, width=30)
        entry.grid(row=i, column=1, padx=10, pady=5)
        entry.insert(0, selected_data[i])
        entries.append(entry)

    # 保存修改的函数
    def save_modifications():
        global data
        try:
            # 获取输入框中的数据
            modified_data = [entry.get() for entry in entries]

            # 验证开始时间和结束时间
            start_time = modified_data[3]
            end_time = modified_data[4]

            if not validate_time(start_time) or not validate_time(end_time):
                messagebox.showerror("错误", "时间格式不正确，请输入有效的24小时制时间（如14:30）")
                return

            # 使用线程锁
            with table_lock:
                # 更新数据
                for i, row in enumerate(data):
                    if f"{row[0]}-{row[1]}-{row[3]}" == selection:
                        data[i] = modified_data
                        break

                # 写入 Excel 文件
                write_to_excel(data)

                # 更新表格
                update_table(data)

                # 更新下拉列表
                movie_drop_down['values'] = [f"{row[0]}-{row[1]}-{row[3]}" for row in data]
                movie_drop_down.set('')  # 清空当前选择的电影

            # 关闭修改窗口
            modify_window.destroy()

            messagebox.showinfo("信息", "电影信息已成功修改并保存！")
        except Exception as e:
            messagebox.showerror("错误", f"保存修改时发生错误: {e}")
            write_error_log(e)

    # 验证时间格式的函数
    def validate_time(time_str):
        try:
            hours, minutes = map(int, time_str.split(':'))
            if 0 <= hours < 24 and 0 <= minutes < 60:
                return True
            return False
        except ValueError:
            return False

    # 保存按钮
    save_button = tk.Button(modify_window, text="保存修改", command=save_modifications)
    save_button.grid(row=len(labels), column=0, columnspan=2, pady=20)
        

def play_audio(filepath):
    try:
        # 确保只初始化一次pygame，避免在多个线程中重复初始化
        if not pygame.get_init():
            pygame.init()
        if not pygame.mixer.get_init():
            pygame.mixer.init()
        
        # 加载音频文件
        sound = pygame.mixer.Sound(filepath)
        # 播放音频
        sound.play()
        while pygame.mixer.get_busy():  # 如果音频正在播放，这将为 True
            pygame.time.Clock().tick(10)  # 限制循环速度，避免占用太多 CPU
    except pygame.error as e:
        write_error_log(f"音频错误：{e}")
    except Exception as e:
        write_error_log(f"播放音频时发生错误：{e}")

# 播放手动广播的函数
def search_data():
    import pygame

    # 检测是否有正在播放的广播进程，如有则警告
    try:
        if pygame.mixer.music.get_busy():
            messagebox.showwarning("警告", "请等待当前广播结束！")
            return
    except:
        pass

    try:    # 避免mixer被关闭，检测未初始化而报错
        pygame.mixer.music.stop()
        pygame.quit()
    except: 
        pass

    # 初始化mixer
    pygame.mixer.init()
    
    def play_announcement():
        selection = movie_drop_down.get()
        if selection:
            # 获取下拉列表中选中项的索引
            index = movie_drop_down.current()
            selected_data = data[index]
            print(selected_data)
            # selected_data = ('好东西', '今天', '12月3', '21:05', '23:08', '3号厅')
            film_name = str(selected_data[0])
            day = str(selected_data[1])
            date = str(selected_data[2])
            start_hour, start_minute = str(selected_data[3]).split(sep=':')[0], str(selected_data[3]).split(sep=':')[1]
            end_hour, end_minute = str(selected_data[4]).split(sep = ':')[0], str(selected_data[4]).split(sep = ':')[1]
            hall_number = str(str(selected_data[5]).split(sep = '号厅')[0])
        
            if hall_number in ['1', '3', '4']:
                check_in_counter = 'left'
            elif hall_number in ['2', '5']:
                check_in_counter = 'right'
            else:
                messagebox.showerror("错误", "未找到符合条件的放映厅！")
                return

            list = [os.path.join('material', 'mix', '756.wav'), os.path.join('material', 'template_cn', '1.wav'), 
            os.path.join('material', 'hall_cn', f'{hall_number}.wav'), 
            os.path.join('material', 'hour_cn', f'{start_hour}.wav'), os.path.join('material', 'minute_cn', f'{start_minute}.wav'),
            os.path.join('material', 'template_cn', '2.wav'), os.path.join('material', 'filmname_cn', f'{film_name}.wav'),
            os.path.join('material', 'template_cn', '3.wav'), os.path.join('material', 'gate_cn', f'{check_in_counter}.wav'), 
            os.path.join('material', 'template_cn', '4.wav')] * cycle_time

            # 756.wav                      --  756提示音
            # template_cn\\1.wav           --  各位观众请注意
            
            # hall_cn\\5.wav               --  五号厅
            
            # hour_cn\\17.wav              --  十七点
            # minute_cn\\15.wav            --  十五分
            
            # template_cn\\2.wav           --  播放的电影
            # filmname_cn\\熊出没.wav      --  熊出没
            
            # template_cn\\3.wav           --  现在开始检票入场，请前往
            # gate_cn\\left.wav            --  左侧检票口
            
            # template_cn\\4.wav           --  检票入场，谢谢！

            # 检查list列表中所有需要的语音片段的路径是否正确，以及对应语音包是否都存在，如果不存在，则报错缺失的所有语音包文件路径
            missing_filename = []
                
            # print(list)
                
            print(f'film_name = {film_name}')
            print(f'day = {day}')
            print(f'date = {date}')
            print(f'start_hour = {start_hour}')
            print(f'start_minute = {start_minute}')
            print(f'end_hour = {end_hour}')
            print(f'end_minute = {end_minute}')
            print(f'hall_number = {hall_number}')

            for wav_file in list:
                if os.path.exists(wav_file) == False:
                    missing_filename.append(wav_file)
            if missing_filename != []:
                # for missing_file in missing_filename:
                messagebox.showwarning("Warning", "Missing wav files!\nNo matching wav files for\n" + str(missing_filename))
                write_error_log("Missing wav files!\nNo matching wav files for\n" + str(missing_filename))
                return



            # list检索完毕，开始生成音频
            import pygame
            
            # 初始化pygame的子系统
            pygame.init()
            
            try:
                pygame.event.get()  # 清空事件队列
            except:
                pass
                
            combined = AudioSegment.empty()  # 初始化 combined 变量


            # 引入进度条窗口
            root = tk.Tk()
            root.title("Importing voice packs")
            root.geometry("500x120")

            # 创建进度条
            progress_bar = ttk.Progressbar(root, orient="horizontal", length=300, mode="determinate")
            progress_bar.pack(pady=20, padx=20, fill="both", expand=True)

            # 创建一个标签控件
            label = tk.Label(root, text="Importing voice packs", font=("Times New Roman", 15))
            label.pack()

            # 循环执行 len(list) 次，也就是要合成 len(list) 个语音片段
            i = 0
            for wav_file in list:
                i += 1
                try:
                    # 读取每个音频文件
                    sound = AudioSegment.from_wav(wav_file)
                    # 将音频文件添加到合并片段中
                    combined += sound
                except Exception as error_message:
                    messagebox.showerror("Error", f"Failed to load audio file: {wav_file}\nError message: {error_message}")
                    write_error_log(f"Failed to load audio file: {wav_file}\nError message: {error_message}")
                    continue

                update_progress(i, len(list), progress_bar, label, list[i - 1])  # 更新进度条和标签
                root.update()  # 刷新主窗口，以便更新进度条和标签

            # 循环完成后，稍作延时，然后销毁窗口
            import time
            time.sleep(0.5)
            root.destroy()
            

            # 检查文件夹是否存在
            if not os.path.exists("output"):
                # 如果文件夹不存在，则创建文件夹
                os.makedirs("output")

            # 将合并的片段导出为wav文件
            try:
                combined.export(os.path.join('output', str(film_name) + '_' + str(date) + '_' + str(start_hour) + '_' + str(start_minute) + '.wav'), format="wav")
            except Exception as error_message:
                # 如果发生异常，捕获异常信息并显示错误消息框
                file_path = os.path.join('output', str(film_name) + '_' + str(date) + '_' + str(start_hour) + '_' + str(start_minute) + '.wav')
                messagebox.showerror("Error", f'An error occurred when exporting {file_path}.\n' + str(error_message))
                write_error_log(error_message)
                return

            time.sleep(0.5)


            # 播放手动广播进程
            try:
                pygame.mixer.music.load(os.path.join('output', str(film_name) + '_' + str(date) + '_' + str(start_hour) + '_' + str(start_minute) + '.wav'))
                pygame.mixer.music.play()
            except Exception as error_message:
                # 如果发生异常，捕获异常信息并显示错误消息框
                file_path = os.path.join('output', str(film_name) + '_' + str(date) + '_' + str(start_hour) + '_' + str(start_minute) + '.wav')
                messagebox.showerror("Error", f'An error occurred when loading {file_path}.\n' + str(error_message))
                write_error_log(error_message)
                return
            
            time.sleep(0.5)
        
            # pygame.mixer.music.get_busy() 判断是否正在播放音乐，返回1为正在播放
            try:    # 避免mixer被关闭，检测未初始化而报错
                while pygame.mixer.music.get_busy():
                    time.sleep(1)
            except:
                pass
            # time.sleep(50)
            
            try:    # 避免mixer被关闭，检测未初始化而报错
                pygame.mixer.music.stop()
                pygame.quit()
            except: 
                pass
    
        else:
            messagebox.showwarning("警告", "请选择一部电影")

    try:
        # 创建并启动播放音频的线程
        audio_thread = threading.Thread(target=play_announcement)
        audio_thread.start()
    except Exception as error_message:
        # 如果发生异常，捕获异常信息并显示错误消息框
        messagebox.showerror("Error", 'An error occurred when playing combined audio.\n' + str(error_message))
        write_error_log(error_message)
        return

# 用于直接从data.xlsx中读取电影信息的函数
def read_from_excel_and_update():
    global data
    try:
        # 尝试获取锁，设置超时时间为1秒
        if table_lock.acquire(timeout=1):
            try:
                data = read_from_excel()  # 从Excel文件读取数据
                update_table(data)  # 更新表格
                messagebox.showinfo("信息", "电影信息已从data.xlsx文件中读取并更新！")
            finally:
                # 释放锁
                table_lock.release()
    except Exception as e:
        print(f"从Excel读取并更新表格时发生错误: {e}")
        write_error_log(e)
    
    # 获取缺失的电影名称
    check_movie_name()

# 读取并刷新的函数
def refresh_data():
    global data, cinema_name, cinema_address
    # 获取屏幕宽高
    screen_width = root.winfo_screenwidth()
    screen_height = root.winfo_screenheight()

    # 设置进度窗口的大小和位置
    progress_window = tk.Toplevel(root)
    progress_window.title("Loading")
    window_width = 300
    window_height = 100
    x = (screen_width // 2) - (window_width // 2)
    y = (screen_height // 2) - (window_height // 2)
    progress_window.geometry(f"{window_width}x{window_height}+{x}+{y}")

    # 添加标签
    tk.Label(progress_window, text="爬取时间大约需要5~10秒，请稍候……").pack()
    progress_window.update()

    # 启动爬取操作
    new_data = fetch_movie_schedules(url, progress_window)

    # 尝试获取锁，设置超时时间为1秒
    if table_lock.acquire(timeout=1):
        try:
            # 更新数据
            data = new_data
            update_table(data)

            # 更新下拉列表
            movie_drop_down['values'] = [f"{row[0]}-{row[1]}-{row[3]}" for row in data]
            movie_drop_down.set('')  # 清空当前选择的电影

            # 更新电影院名称和地址的标签
            cinema_info_label.config(text=f"电影院：{cinema_name} 影院地址：{cinema_address}")
        finally:
            # 释放锁
            table_lock.release()

    progress_window.destroy()
    if successful:
        messagebox.showinfo("信息", "所有电影信息已成功爬取，并全部刷新！")
    elif not successful:
        messagebox.showwarning("警告", "电影信息爬取失败，请稍后再试！")
    else:
        messagebox.showinfo("信息", "遇到了一些问题！")
    
    # 获取缺失的电影名称
    check_movie_name()

def clear_and_exit():
    global data
    data = []
    update_table(data)
    # 如output文件夹存在，则清空output文件夹中的缓存
    try:
        delete_all_files_in_directory("output")
    except:
        pass
    messagebox.showinfo("信息", "所有缓存已清空并退出程序。")
    root.destroy()

def get_pre_minute_and_cycle_time():
    global pre_minute, cycle_time  # 声明全局变量
    # 创建一个新的 Tkinter 窗口
    settings_window = tk.Tk()
    settings_window.title("设置")

    # 设置窗口大小
    window_width = 300
    window_height = 200
    screen_width = settings_window.winfo_screenwidth()
    screen_height = settings_window.winfo_screenheight()
    x = (screen_width // 2) - (window_width // 2)
    y = (screen_height // 2) - (window_height // 2)
    settings_window.geometry(f"{window_width}x{window_height}+{x}+{y}")

    # 创建一个标签提示用户选择提前检票的分钟数
    tk.Label(settings_window, text="请选择提前检票的分钟数：").pack(pady=5)

    # 创建一个下拉列表供用户选择分钟数
    pre_minute_options = [str(i) for i in range(1, 31)]  # 创建1-30的整数列表
    entry_var = tk.StringVar(value=str(pre_minute))  # 设置默认值
    pre_minute_drop_down = ttk.Combobox(settings_window, textvariable=entry_var, values=pre_minute_options)
    pre_minute_drop_down.pack(pady=5)

    # 创建一个标签提示用户选择重复播放次数
    tk.Label(settings_window, text="请选择重复播放次数：").pack(pady=5)

    # 创建一个下拉列表供用户选择重复播放次数
    cycle_time_options = [str(i) for i in range(1, 11)]  # 创建1-10的整数列表
    cycle_var = tk.StringVar(value=str(cycle_time))  # 设置默认值
    cycle_time_drop_down = ttk.Combobox(settings_window, textvariable=cycle_var, values=cycle_time_options)
    cycle_time_drop_down.pack(pady=5)

    # 创建一个按钮，用户选择完毕后点击确认
    def on_confirm():
        global pre_minute, cycle_time  # 声明全局变量
        try:
            pre_minute = int(entry_var.get())  # 将字符串转换为整数
            cycle_time = int(cycle_var.get())  # 将字符串转换为整数
            settings_window.destroy()  # 关闭小窗口
        except ValueError:
            messagebox.showerror("错误", "请输入有效的分钟数或播放次数")

    confirm_button = tk.Button(settings_window, text="确认", command=on_confirm)
    confirm_button.pack(pady=10)

    # 运行小窗口的事件循环
    settings_window.mainloop()

# 播放自动广播的函数，检测开场时间满足提前检票时间，自动播放检票广播
def check_movies():
    while True:
        # 初始化pygame和pygame.mixer
        if not pygame.get_init():
            pygame.init()
        if not pygame.mixer.get_init():
            pygame.mixer.init()
            
        import time
        current_time = time.strftime('%H:%M', time.localtime(time.time()))
        
        # 尝试获取锁，设置超时时间为1秒
        if table_lock.acquire(timeout=1):
            try:
                for row in data:
                    film_name = str(row[0])
                    day = str(row[1])
                    date = str(row[2])
                    start_hour, start_minute = str(row[3]).split(sep=':')[0], str(row[3]).split(sep=':')[1]
                    end_hour, end_minute = str(row[4]).split(sep=':')[0], str(row[4]).split(sep=':')[1]
                    hall_number = str(str(row[5]).split(sep='号厅')[0])

                    start_min = int(start_hour) * 60 + int(start_minute)
                    current_min = int(current_time.split(':')[0]) * 60 + int(current_time.split(':')[1])
                    delta = start_min - current_min

                    film_key = f"{row[0]}-{row[1]}-{row[3]}"

                    if day == '今天' and 0 <= delta <= pre_minute and film_key not in film_played:
                        print(f'film_key = {film_key}')
                        print(f'film_played = {film_played}')

                        if hall_number in ['1', '3', '4']:
                            check_in_counter = 'left'
                        elif hall_number in ['2', '5']:
                            check_in_counter = 'right'
                        else:
                            # 使用主线程执行 messagebox.showerror
                            root.after(0, lambda: messagebox.showerror("错误", "未找到符合条件的放映厅！"))
                            return

                        list = [os.path.join('material', 'mix', '756.wav'), os.path.join('material', 'template_cn', '1.wav'), 
                                os.path.join('material', 'hall_cn', f'{hall_number}.wav'), 
                                os.path.join('material', 'hour_cn', f'{start_hour}.wav'), os.path.join('material', 'minute_cn', f'{start_minute}.wav'),
                                os.path.join('material', 'template_cn', '2.wav'), os.path.join('material', 'filmname_cn', f'{film_name}.wav'),
                                os.path.join('material', 'template_cn', '3.wav'), os.path.join('material', 'gate_cn', f'{check_in_counter}.wav'), 
                                os.path.join('material', 'template_cn', '4.wav')] * cycle_time

                        missing_filename = []

                        print(f'film_name = {film_name}')
                        print(f'day = {day}')
                        print(f'date = {date}')
                        print(f'start_hour = {start_hour}')
                        print(f'start_minute = {start_minute}')
                        print(f'end_hour = {end_hour}')
                        print(f'end_minute = {end_minute}')
                        print(f'hall_number = {hall_number}')

                        for wav_file in list:
                            if not os.path.exists(wav_file):
                                missing_filename.append(wav_file)
                        if missing_filename:
                            # 使用主线程执行 messagebox.showwarning
                            root.after(0, lambda: messagebox.showwarning("Warning", "Missing wav files!\nNo matching wav files for\n" + str(missing_filename)))
                            write_error_log("Missing wav files!\nNo matching wav files for\n" + str(missing_filename))
                            return

                        film_played.append(film_key)  # 添加到已播放列表
                        print(f'film_played = {film_played}')

                        combined = AudioSegment.empty()  # 初始化 combined 变量

                        for wav_file in list:
                            try:
                                sound = AudioSegment.from_wav(wav_file)
                                combined += sound
                            except Exception as error_message:
                                root.after(0, lambda: messagebox.showerror("Error", f"Failed to load audio file: {wav_file}\nError message: {error_message}"))
                                write_error_log(f"Failed to load audio file: {wav_file}\nError message: {error_message}")
                                continue

                        if not os.path.exists("output"):
                            os.makedirs("output")

                        try:
                            combined.export(os.path.join('output', str(film_name) + '_' + str(date) + '_' + str(start_hour) + '_' + str(start_minute) + '.wav'), format="wav")
                        except Exception as error_message:
                            root.after(0, lambda: messagebox.showerror("Error", f'An error occurred when exporting {os.path.join("output", str(film_name) + "_" + str(date) + "_" + str(start_hour) + "_" + str(start_minute) + ".wav")}.\n' + str(error_message)))
                            write_error_log(error_message)
                            return

                        try:
                            pygame.mixer.music.load(os.path.join('output', str(film_name) + '_' + str(date) + '_' + str(start_hour) + '_' + str(start_minute) + '.wav'))
                            pygame.mixer.music.play()
                        except Exception as error_message:
                            root.after(0, lambda: messagebox.showerror("Error", f'An error occurred when loading {os.path.join("output", str(film_name) + "_" + str(date) + "_" + str(start_hour) + "_" + str(start_minute) + ".wav")}.\n' + str(error_message)))
                            write_error_log(error_message)
                            return

                        try:
                            while pygame.mixer.music.get_busy():
                                time.sleep(1)
                        except:
                            pass

                        try:    # 避免mixer被关闭，检测未初始化而报错
                            pygame.mixer.music.stop()
                            pygame.quit()
                        except: 
                            pass
            finally:
                # 释放锁
                table_lock.release()

        time.sleep(5)  # 每隔5秒检查一次

# 创建一个函数来停止所有音频
def stop_all_audio():
    import pygame
    try:
        # 确保只初始化一次pygame，避免在多个线程中重复初始化
        if not pygame.get_init():
            pygame.init()
        if not pygame.mixer.get_init():
            pygame.mixer.init()
        
        # 检查是否有音频正在播放
        if pygame.mixer.music.get_busy():
            pygame.mixer.music.stop()
        
        # 可选：延迟一点时间，确保音频已停止
        time.sleep(0.5)
        
    except pygame.error as e:
        print(f"停止音频播放时发生错误: {e}")
    finally:
        # 播放完成后退出pygame
        pygame.quit()
        
# 定义一个函数来检查正在播放的电影
import threading

# 创建一个线程锁
table_lock = threading.Lock()

# 检查正在播放的电影的函数
def check_playing_movies():
    while True:
        # 获取当前系统时间
        current_time = datetime.now().strftime('%H:%M')
        current_time_obj = datetime.strptime(current_time, '%H:%M')  # 将当前时间转换为datetime对象
        
        # 尝试获取锁，设置超时时间为1秒
        if table_lock.acquire(timeout=1):
            try:
                # 遍历表格中的所有行
                for item in table.get_children():
                    try:
                        # 获取每一行的数据
                        row_data = table.item(item, 'values')
                        
                        date = row_data[1]
                        start_time = row_data[3]
                        end_time = row_data[4]
                        
                        # 将开始时间和结束时间转换为datetime对象
                        start_time_obj = datetime.strptime(start_time, '%H:%M')
                        end_time_obj = datetime.strptime(end_time, '%H:%M')
                        
                        # 计算当前时间与结束时间的差值（以分钟为单位）
                        time_difference_end = (end_time_obj - current_time_obj).total_seconds() / 60
                        time_difference_start = (start_time_obj - current_time_obj).total_seconds() / 60  # 新增

                        # 判断当前时间是否在开始时间和结束时间之间
                        if start_time <= current_time <= end_time and date == '今天':
                            # 如果正在播放，设置背景为绿色lightgreen
                            table.tag_configure('playing', background='lightgreen')
                            table.item(item, tags=('playing',))
                            
                            # 如果电影在10分钟内结束，设置背景为浅红色#F08080
                            if 0 <= time_difference_end <= 10:
                                table.tag_configure('ending_soon', background='#F08080')
                                table.item(item, tags=('ending_soon',))
                        # 新增：判断是否在10分钟内即将开始
                        elif 0 <= time_difference_start <= 10 and date == '今天':
                            table.tag_configure('upcoming', background='yellow')
                            table.item(item, tags=('upcoming',))
                        else:
                            # 如果不在播放，清除背景色
                            table.item(item, tags=('',))
                    except Exception as e:
                        print(f"检查播放电影时发生错误: {e}")
                        write_error_log(e)
            finally:
                # 释放锁
                table_lock.release()
        
        # 每5秒检查一次
        time.sleep(5)

# 修改电影信息的函数
def modify_movie_info():
    global data
    selection = movie_drop_down.get()
    if not selection:
        messagebox.showwarning("警告", "请选择一部电影")
        return

    # 获取选中的电影信息
    selected_data = None
    for row in data:
        if f"{row[0]}-{row[1]}-{row[3]}" == selection:
            selected_data = row
            break

    if not selected_data:
        messagebox.showerror("错误", "未找到选中的电影信息")
        return

    # 弹出修改窗口
    modify_window = tk.Toplevel(root)
    modify_window.title("修改电影信息")

    # 获取屏幕宽高
    screen_width = modify_window.winfo_screenwidth()
    screen_height = modify_window.winfo_screenheight()

    # 设置窗口大小
    window_width = 350
    window_height = 300

    # 计算居中位置
    x = (screen_width // 2) - (window_width // 2)
    y = (screen_height // 2) - (window_height // 2)

    # 设置窗口大小及位置
    modify_window.geometry(f"{window_width}x{window_height}+{x}+{y}")

    # 创建输入框和标签
    labels = ["电影名称", "日期", "日期详情", "开始时间", "结束时间", "放映厅"]
    entries = []
    for i, label_text in enumerate(labels):
        tk.Label(modify_window, text=label_text).grid(row=i, column=0, padx=10, pady=5)
        entry = tk.Entry(modify_window, width=30)
        entry.grid(row=i, column=1, padx=10, pady=5)
        entry.insert(0, selected_data[i])
        entries.append(entry)

    # 保存修改的函数
    def save_modifications():
        global data
        try:
            # 获取输入框中的数据
            modified_data = [entry.get() for entry in entries]

            # 验证开始时间和结束时间
            start_time = modified_data[3]
            end_time = modified_data[4]

            if not validate_time(start_time) or not validate_time(end_time):
                messagebox.showerror("错误", "时间格式不正确，请输入有效的24小时制时间（如14:30）")
                return

            # 尝试获取锁，设置超时时间为1秒
            if table_lock.acquire(timeout=1):
                try:
                    # 更新数据
                    for i, row in enumerate(data):
                        if f"{row[0]}-{row[1]}-{row[3]}" == selection:
                            data[i] = modified_data
                            break

                    # 写入 Excel 文件
                    write_to_excel(data)

                    # 更新表格
                    update_table(data)

                    # 更新下拉列表
                    movie_drop_down['values'] = [f"{row[0]}-{row[1]}-{row[3]}" for row in data]
                    movie_drop_down.set('')  # 清空当前选择的电影
                finally:
                    # 释放锁
                    table_lock.release()

            # 关闭修改窗口
            modify_window.destroy()

            messagebox.showinfo("信息", "电影信息已成功修改并保存！")
        except Exception as e:
            messagebox.showerror("错误", f"保存修改时发生错误: {e}")
            write_error_log(e)

    # 验证时间格式的函数
    def validate_time(time_str):
        try:
            hours, minutes = map(int, time_str.split(':'))
            if 0 <= hours < 24 and 0 <= minutes < 60:
                return True
            return False
        except ValueError:
            return False

    # 保存按钮
    save_button = tk.Button(modify_window, text="保存修改", command=save_modifications)
    save_button.grid(row=len(labels), column=0, columnspan=2, pady=20)
        
# 更新当前时间的函数
def update_time():
    current_time = time.strftime('%H:%M:%S', time.localtime(time.time()))
    time_label.config(text=current_time)
    # 每隔1秒更新一次时间
    root.after(1000, update_time)
        
# 主程序
if __name__ == '__main__':
    # 初始化null值
    cinema_name = "null"
    cinema_address = "null"

    # 初始化pygame和pygame.mixer
    pygame.init()
    pygame.mixer.init()
    
    convert_to_stereo("material")     # 将material文件夹下所有wav音频文件转换为双声道
    
    url = url   # 要爬取的影院网页，已在info.txt文件中读取
    
    data = read_from_excel()  # 从Excel文件读取数据，而不是在启动时爬取数据
    
    # 先弹出小窗口让用户设置提前检票分钟数和重复播放次数
    get_pre_minute_and_cycle_time()
    
    # 获取缺失的电影名称
    check_movie_name()

    # 创建主窗口并使其居中
    root = tk.Tk()
    root.title("电影院自动广播系统")
    
    # 设置窗口大小
    window_width = 650
    window_height = 600
    
    # 获取屏幕宽高
    screen_width = root.winfo_screenwidth()
    screen_height = root.winfo_screenheight()
    
    # 计算居中位置
    x = (screen_width // 2) - (window_width // 2)
    y = (screen_height // 2) - (window_height // 2)
    
    # 设置窗口大小及位置
    root.geometry(f"{window_width}x{window_height}+{x}+{y}")
    
    # 表格显示区域样式设置
    style = ttk.Style()
    style.configure("Treeview", font=("Microsoft YaHei", 12, "bold"), rowheight=30)  # 设定字体，字号，加粗，行高

    # 创建表格
    table = ttk.Treeview(root, columns=("Filmname", "Date name", "Date day", "Start time", "End time", "Hall No."), show="headings")
    
    # 设置每列的标题和宽度
    table.heading("Filmname", text="电影名称")
    table.column("Filmname", width=240, anchor=tk.CENTER)  # 设置电影名称列的宽度为240像素，文本居中显示

    table.heading("Date name", text="日期")
    table.column("Date name", width=50, anchor=tk.CENTER)  # 设置日期列的宽度为50像素，文本居中显示

    table.heading("Date day", text="日期详情")
    table.column("Date day", width=80, anchor=tk.CENTER)  # 设置日期详情列的宽度为80像素，文本居中显示

    table.heading("Start time", text="开始时间")
    table.column("Start time", width=80, anchor=tk.CENTER)  # 设置开始时间列的宽度为80像素，文本居中显示

    table.heading("End time", text="结束时间")
    table.column("End time", width=80, anchor=tk.CENTER)  # 设置结束时间列的宽度为80像素，文本居中显示

    table.heading("Hall No.", text="放映厅")
    table.column("Hall No.", width=120, anchor=tk.CENTER)  # 设置放映厅列的宽度为120像素，文本居中显示

    table.pack(fill='both', expand=True)

    # 绑定表格的选中事件
    def on_table_select(event):
        # 获取选中的行
        selected_item = table.selection()
        if selected_item:
            # 获取选中行的数据
            selected_data = table.item(selected_item, 'values')
            # 更新下拉列表的值
            movie_drop_down.set(f"{selected_data[0]}-{selected_data[1]}-{selected_data[3]}")

    # 绑定事件
    table.bind('<<TreeviewSelect>>', on_table_select)
    
    # 创建右键菜单
    def create_context_menu(event):
        # 获取鼠标点击的位置
        row_id = table.identify_row(event.y)
        if row_id:
            # 选中该行
            table.selection_set(row_id)
            table.focus(row_id)
        
            # 获取选中行的数据
            selected_data = table.item(row_id, 'values')
            movie_drop_down.set(f"{selected_data[0]}-{selected_data[1]}-{selected_data[3]}")
        
            # 创建右键菜单
            context_menu = tk.Menu(root, tearoff=0)
            context_menu.add_command(label="播放广播", command=search_data)
            context_menu.add_command(label="修改电影信息", command=modify_movie_info)
        
            # 显示右键菜单
            context_menu.post(event.x_root, event.y_root)
        else:
            print("未选中任何行")
    
    # 绑定右键事件
    table.bind("<Button-3>", create_context_menu)

    # 更新表格
    update_table(data)

    # 创建下拉列表
    movie_selection = tk.StringVar(root)
    
    # 第1行
    first_frame = tk.Frame(root)
    first_frame.pack(fill='x')
    tk.Label(first_frame, text="选择电影:").pack(side=tk.LEFT, padx=10, pady=10)
    movie_drop_down = ttk.Combobox(first_frame, textvariable=movie_selection, values=[f"{row[0]}-{row[1]}-{row[3]}" for row in data], width=30)
    movie_drop_down.pack(side=tk.LEFT, padx=10, pady=10)

    # 按钮行
    button_frame = tk.Frame(root)
    button_frame.pack(fill='x')
    # 创建按钮
    display_button = tk.Button(button_frame, text="播放广播", command=search_data)
    display_button.pack(side=tk.LEFT, padx=10, pady=10)
    stop_button = tk.Button(button_frame, text="停止播放", command=stop_all_audio)
    stop_button.pack(side=tk.LEFT, padx=5, pady=5)
    
    delete_button = tk.Button(button_frame, text="清空缓存", command=lambda: delete_all_files_in_directory("output") and messagebox.showinfo("信息", "缓存已清空"))
    delete_button.pack(side=tk.LEFT, padx=5, pady=5)
    
    clear_button = tk.Button(button_frame, text="清空并退出", command=clear_and_exit)
    clear_button.pack(side=tk.RIGHT, padx=10, pady=10)
    refresh_button = tk.Button(button_frame, text="读取并刷新", command=refresh_data)
    refresh_button.pack(side=tk.RIGHT, padx=10, pady=10)
    
    # 从data.xlsx读取电影信息的按钮
    read_excel_button = tk.Button(button_frame, text="从Excel读取", command=read_from_excel_and_update)
    read_excel_button.pack(side=tk.RIGHT, padx=10, pady=10)
    
    # 修改电影信息的按钮
    modify_button = tk.Button(button_frame, text="修改电影信息", command=modify_movie_info)
    modify_button.pack(side=tk.RIGHT, padx=10, pady=10)
    
    # 创建一个标签，用于显示当前时间
    time_label = tk.Label(first_frame, text="", font=("Times New Roman", 30))
    time_label.pack(side=tk.RIGHT, padx=10, pady=10)
    
    # 第2行
    second_frame = tk.Frame(root)
    second_frame.pack(fill='x')
    tk.Label(second_frame, text=f"已设置提前 {pre_minute} 分钟检票，每个广播循环播放 {cycle_time} 次。").pack(side=tk.LEFT, padx=10, pady=1)
    tk.Label(second_frame, text=f"*黄色-即将开场*  *绿色-正在播放*  *红色-即将散场*").pack(side=tk.RIGHT, padx=10, pady=1)
    
    # 第3行，用来显示电影院名称和地址
    cinema_frame = tk.Frame(root)
    cinema_frame.pack(fill='x')
    cinema_info_label = tk.Label(cinema_frame, text=f"电影院：{cinema_name} 影院地址：{cinema_address}")
    cinema_info_label.pack(side=tk.LEFT, padx=10, pady=1)
    
    # 第4行
    fourth_frame = tk.Frame(root)
    fourth_frame.pack(fill='x')
    tk.Label(fourth_frame, text=f"软件版权归属于：吴瀚庆    版本号：{version_code}    欢迎联系软件作者：whq20050121").pack(side=tk.LEFT, padx=10, pady=1)
    
    # 第5行
    # fifth_frame = tk.Frame(root)
    # fifth_frame.pack(fill='x')
    # tk.Label(fifth_frame, text=f"请避免过于频繁停止播放，请定期清空缓存，以便清理output文件夹下已合成的音频。").pack(side=tk.LEFT, padx=10, pady=1)

    # 启动定时检测线程
    threading.Thread(target=check_movies, daemon=True).start()

    # 启动检查正在播放电影的线程
    threading.Thread(target=check_playing_movies, daemon=True).start()

    # 启动时间更新
    update_time()

    root.mainloop()
